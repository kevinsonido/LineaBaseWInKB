import ctypes
import io
from ipaddress import ip_address
import os
import codecs
import getpass
from pydoc import synopsis
import re
from sys import path
import sys
from winreg import *
import platform,getpass
import socket
from datetime import date, datetime
import winreg


from django.http import response, FileResponse
from django.shortcuts import HttpResponse, redirect, render

#para mjs en login/registro
from django.contrib import messages
from django.contrib.auth import authenticate,login,logout
from django.contrib.auth.decorators import login_required

#reporte#

from .static.LineaBaseApp.assets.Modulos.AccountPolicies import MinimumPasswordLength, temp
from .static.LineaBaseApp.assets.Modulos.Observaciones import *
from .models import *
from .models import tb_plugin, tb_estadolb,HistorialEscaneo, tb_scan
from .forms import CrearUsuarioForm
import csv
from openpyxl import Workbook
from openpyxl.styles import Font


# Create your views here.3
@login_required(login_url='login') 
def home(request):
     #accountPolicies
     MiniPassAge = MinimumPasswordAge()
     Minpasslength = MinimumPasswordLength()
     MaxPassAge = MaximumPasswordAge()
     PassComplex = PasswordComplexity()
     ClTPass = ClearTextPassword()
     LockDur = LockoutDuration()
       # FIN accountPolicies 
     # Local Policies
     Acccomputer=SeNetworkLogonRight()
     denyaccess= SeDenyNetworkLogonRight()
     denylog = SeDenyInteractiveLogonRight()
     allowlog = SeRemoteInteractiveLogonRight()
   
     #opciones de seguridad
     dontdisplay = DontDisplayLastUserName()
     passexp=PasswordExpiryWARNING()
     smartcard= ScRemoveOption()
     digitaly = RequireSecuritySignature()
     nullsesion=NullSessionPipes()
     restricnull=restrictnullsessaccess()
     
     controlFallido= Fallidos()
     ControlAcept = Passed()
     cart =0
     cart2=0
     for e in range(1,controlFallido+1):
          cart = e
     print(cart)

     for a in range(1,ControlAcept+1):
          cart2=a
     print(cart2)
          
          
          
     return render(request,"ProyectoLineaBase\Home.html",{"Fallidos":cart,"Aceptados":cart2,"minpass":MiniPassAge,
                                                               "minpassle":Minpasslength,"maxpassage":MaxPassAge,"passcomplex":PassComplex,
                                                               "cltpass":ClTPass,"lockdur":LockDur,"acccompu":Acccomputer,"denylog":denylog,
                                                               "denyacce":denyaccess,"allowlo":allowlog,
                                                               "dontdisplays":dontdisplay,"passexperi":passexp,"smartcar":smartcard, 
                                                               "digi":digitaly,"nullse":nullsesion,"restrictnull":restricnull,})

#request es una solicitud
@login_required(login_url='login') 
def lineabase(request):
     #informacion del sistema#
     fechaactual = datetime.now() 
     System_Name = platform.uname()[1]
     #Os_name=platform.system(),platform.release(),platform.win32_ver()[2] 
     Os_name=platform.system(),platform.release()
     Osversion = platform.version()
     System_Name2 = platform.uname()[2]
     user_name= platform.uname()[1],getpass.getuser()  
     # Ip del equipo.
     IP_host =socket.gethostbyname(socket.gethostname())
     #puerto
     sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
     sock.bind(('0.0.0.0', 0))
     
   #accountPolicies
     MiniPassAge = MinimumPasswordAge()
     Minpasslength = MinimumPasswordLength()
     MaxPassAge = MaximumPasswordAge()
     PassComplex = PasswordComplexity()
     ClTPass = ClearTextPassword()
     LockDur = LockoutDuration()
       # FIN accountPolicies 
     # Local Policies
     Acccomputer=SeNetworkLogonRight()
     denyaccess= SeDenyNetworkLogonRight()
     denylog = SeDenyInteractiveLogonRight()
     allowlog = SeRemoteInteractiveLogonRight()
   
     #opciones de seguridad
     dontdisplay = DontDisplayLastUserName()
     passexp=PasswordExpiryWARNING()
     smartcard= ScRemoveOption()
     digitaly = RequireSecuritySignature()
     nullsesion=NullSessionPipes()
     restricnull=restrictnullsessaccess()
     #opciones de seguridad
     dontdisplay = DontDisplayLastUserName()
     passexp=PasswordExpiryWARNING()
     smartcard= ScRemoveOption()
     digitaly = RequireSecuritySignature()
     nullsesion=NullSessionPipes()
     restricnull=restrictnullsessaccess()
     ##########	19.1 Control Panel
     get_user=get_user_sid()
     print(get_user)    
     scren= ScreenSaveActive()
     save = SCRNSAVE()
     screen =ScreenSaverIsSecure()
     screentime=ScreenSaveTimeOut()
     always= AlwaysInstallElevated()
     prevent = PreventCodecDownload() 
     #contador
     controlFallido= Fallidos()
     ControlAcept = Passed()
     cart =0
     ca2=0
     for e in range(1,controlFallido+1):
          cart = e
          print(cart)

     for a in range(1,ControlAcept+1):
          ca2=a
          print(ca2)
     # Fin contador
     
     #insertar Tabla Host
     host = tb_host.objects.get(pk=1)
     d1= tb_scan.objects.get(idscan=1)
     host.scan_id=d1
     #host.save()
     infSys=tb_host(idhost=host,iphost=IP_host,fqdnhost=System_Name,
                    starthost=System_Name,endhost=System_Name, oshost=Os_name,scan_id=d1)
     infSys.save()
     print("Se inserto correctamente la TB_Host")
  
    
     #lb = HistorialEscaneo.objects.all()
     
     #mostra en pantalla 
     lb= tb_lb_output.objects.all()
     return render(request,"ProyectoLineaBase\lineabase.html",{"historicos": lb,"fecha_actual":fechaactual,
                                                               "system_name": System_Name,"Os_name":Os_name,"Os_version":Osversion,
                                                               "system_name2":System_Name2, "usuariologueado":user_name ,"minpass":MiniPassAge,
                                                               "minpassle":Minpasslength,"maxpassage":MaxPassAge,"passcomplex":PassComplex,
                                                               "cltpass":ClTPass,"lockdur":LockDur,"acccompu":Acccomputer,"denylog":denylog,
                                                               "denyacce":denyaccess,"allowlo":allowlog,
                                                               "dontdisplays":dontdisplay,"passexperi":passexp,"smartcar":smartcard, 
                                                               "digi":digitaly,"nullse":nullsesion,"restrictnull":restricnull,
                                                               "scren":scren,"save":save,"screen":screen,"screentime":screentime,"always":always,"prevent":prevent,
                                                               "Fallidos":cart,"Aceptados":a,
                                                               "IP":IP_host,"puerto":sock.getsockname()[1]})


def controlesLineaBase(request):
     fechaactual = datetime.now() 
     nombresistema = platform.uname()[1]
     #Os_name=platform.system(),platform.release(),platform.win32_ver()[2] 
     Os_name=platform.system(),platform.release()
     Osversion = platform.version()
     System_Name2 = platform.uname()[2]
     user_name= platform.uname()[1],getpass.getuser()  
     # Ip del equipo.
     IP_host =socket.gethostbyname(socket.gethostname())
     #puerto
     sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
     sock.bind(('0.0.0.0', 0))
     
     MiniPassAge = MinimumPasswordAge()
     Minpasslength = MinimumPasswordLength()
     MaxPassAge = MaximumPasswordAge()
     PassComplex = PasswordComplexity()
     ClTPass = ClearTextPassword()
     LockDur = LockoutDuration()
       # FIN accountPolicies 
     # Local Policies
     Acccomputer=SeNetworkLogonRight()
     denyaccess= SeDenyNetworkLogonRight()
     denylog = SeDenyInteractiveLogonRight()
     allowlog = SeRemoteInteractiveLogonRight()
   
     #opciones de seguridad
     dontdisplay = DontDisplayLastUserName()
     passexp=PasswordExpiryWARNING()
     smartcard= ScRemoveOption()
     digitaly = RequireSecuritySignature()
     nullsesion=NullSessionPipes()
     restricnull=restrictnullsessaccess()
     ##########	19.1 Control Panel
     get_user=get_user_sid()
     print(get_user)    
     scren= ScreenSaveActive()
     save = SCRNSAVE()
     screen =ScreenSaverIsSecure()
     screentime=ScreenSaveTimeOut()
     always= AlwaysInstallElevated()
     prevent = PreventCodecDownload()
     
     #nombre de folder
     #insertar Tabla Host
     fol = tb_folder.objects.get(pk=1)
     d1= tb_scan.objects.get(idscan=1)
     fol.scan_id=d1
     
     #fol= tb_folder._meta.pk.id
     #host.save()
     
     fecha= datetime.today()
     fechaa = fecha.strftime("%d-%b-%y")
     noa= 'EV'+ fechaa
     folder = tb_folder(tipofolder='Normal', nombrefolder=noa)
     li=[]

     scan = tb_scan(tiposcan='Demanda',nombrescan= noa)
     tb_folder.save(folder)
     tb_scan.save(scan)
     #nombre de folder
     
     #guardar registro de escaneo
     scanrun= tb_scan_run(starscan=fechaa,endscan=fechaa)
     
     tb_scan_run.save(scanrun)

     
     if MiniPassAge == "Failed":
          t1 ="Account Policies \ Password Policy \ Minimum password age"
          sub="(L1) Ensure 'Minimum password age' is set to '1 or more day(s)"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="This policy setting determines the number of days that you must use a password before you can change it. The range of values for this policy setting is between 1 and 999 days. (You may also set the value to 0 to allow immediate password changes.) The default value for this setting is 0 days."
          d2="To establish the recommended configuration via GP, set the following UI path to 1 or more day(s): Computer Configuration \ Policies \ Windows Settings \ Security Settings \ Account Policies \ Password Policy\Minimum password age "
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=MiniPassAge,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=MiniPassAge)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=MiniPassAge,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=MiniPassAge)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if Minpasslength == "Failed":
          t1="Account Policies \ Password Policy \ Minimum password age"
          sub="(L1) Ensure 'Minimum password age' is set to '1 or more day(s)"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="This policy setting determines the number of days that you must use a password before you can change it. The range of values for this policy setting is between 1 and 999 days. (You may also set the value to 0 to allow immediate password changes.) The default value for this setting is 0 days."
          d2= "To establish the recommended configuration via GP, set the following UI path to 1 or more day(s): Computer Configuration \ Policies \ Windows Setting s \ Security Settings \ Account Polici"
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=Minpasslength,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=Minpasslength)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=Minpasslength,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=Minpasslength)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if MaxPassAge == "Failed":
          t1="Local Policies \ Security Options \ Domain member \ Domain member: Maximum machine account password age "
          sub="(L1) Ensure 'Domain member: Maximum machine account password age' is set to '30 or fewer days, but not 0'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="This policy setting determines the maximum allowable age for a computer account password. By default, domain members automatically change their domain passwords every 30 days. If you increase this interval significantly so that the computers no longer change their passwords, an attacker would have more time to undertake a brute force attack against one of the computer accounts."
          d2="To establish the recommended configuration via GP, set the following UI path to 30 or fewer days, but not 0: Computer Configuration\Policies\Windows Settings \ Security Settings \ Local Policies \ Security Options \ Domain member: Maximum machine account password age"
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=MaxPassAge,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=MaxPassAge)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=MaxPassAge,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=MaxPassAge)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if PassComplex == "Failed":
          t1="Administrative Templates (Computer) \ LAPS \ Password Settings: Password Complexity"
          sub="(L1) Ensure 'Password Settings: Password Complexity' is set to 'Enabled: Large letters + small letters + numbers + special characters'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="In May 2015, Microsoft released the Local Administrator Password Solution (LAPS) tool, which is free and supported software that allows an organization to automatically set randomized and unique local Administrator account passwords on domain-attached workstations and member servers"
          d2="To establish the recommended configuration via GP, set the following UI path to Enabled, and configure the Password Complexity option to Large letters + small letters + numbers + special characters: Computer Configuration \ Policies \ Administrative Templates \ LAPS \ Password Settings <br /> Note: This Group Policy path does not exist by default. An additional Group Policy template (AdmPwd.admx/adml) is required - it is included with Microsoft Local Administrator Password Solution (LAPS). "
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=PassComplex,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=PassComplex)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=PassComplex,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=PassComplex)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if ClTPass == "Failed":
          t1="Account Policies \ Password Policy \  Store passwords using reversible encryption"
          sub="(L1) Ensure 'Store passwords using reversible encryption' is set to 'Disabled'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="This policy setting determines whether the operating system stores passwords in a way that uses reversible encryption, which provides support for application protocols that require knowledge of the user's password for authentication purposes. Passwords that are stored with reversible encryption are essentially the same as plaintext versions of the passwords."
          d2="To establish the recommended configuration via GP, set the following UI path to Disabled: Computer Configuration \ Policies Windows Settings \ Security Settings \ Account Policies \ Password Policy \ Store passwords using reversible encryption"
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=ClTPass,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=ClTPass)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=ClTPass,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=ClTPass)
          tb_estadolb.save(es)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if LockDur == "Failed":
          t1="Account Policies \ Account Lockout Policy \ Account lockout duration"
          sub="(L1) Ensure 'Account lockout duration' is set to '15 or more minute(s)'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="This policy setting determines the length of time that must pass before a locked account is unlocked and a user can try to log on again. The setting does this by specifying the number of minutes a locked out account will remain unavailable. If the value for this policy setting is configured to 0, locked out accounts will remain locked out until an administrator manually unlocks them."
          d2="To establish the recommended configuration via GP, set the following UI path to 15 or more minute(s): Computer Configuration \ Policies \ Windows Settings \ Security Settings \ Account Policies \ Account Lockout Policy \ Account lockout duration"
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=LockDur,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=LockDur)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=LockDur,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=LockDur)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)       
     if Acccomputer == "Failed":
          t1="Local Policies \ User Rights Assignment \ Access this computer from the network"
          sub="(L1) Ensure 'Access this computer from the network' is set to 'Administrators'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="This policy setting allows other users on the network to connect to the computer and is required by various network protocols that include Server Message Block (SMB)-based protocols, NetBIOS, Common Internet File System (CIFS), and Component Object Model Plus (COM+)."
          d2="To establish the recommended configuration via GP, set the following UI path to Administrators: Computer Configuration \ Policies \ Windows Settings \ Security Settings \ Local Policies \ User Rights Assignment \ Access this computer from the network"

          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=Acccomputer,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=Acccomputer)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=Acccomputer,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=Acccomputer)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if denyaccess == "Failed":
          t1="Local Policies \ User Rights Assignment \ Deny access to this computer from the network"
          sub="(L1) Ensure 'Deny access to this computer from the network' to include 'Guests, Local account'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="This policy setting prohibits users from connecting to a computer from across the network, which would allow users to access and potentially modify data remotely. In high security environments, there should be no need for remote users to access data on a computer. Instead, file sharing should be accomplished through the use of network servers."
          d2="To establish the recommended configuration via GP, set the following UI path to include Guests, Local account: Computer Configuration \ Policies \ Windows Settings \ Security Settings \ Local Policies \ User Rights Assignment\Deny access to this computer from the network"
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=denyaccess,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=denyaccess)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=denyaccess,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=denyaccess)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if denylog == "Failed":
          t1="Local Policies \ User Rights Assignment \ Deny log on locally"
          sub="(L1) Ensure 'Deny log on locally' to include 'Guests'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="This security setting determines which users are prevented from logging on at the computer. This policy setting supersedes the Allow log on locally policy setting if an account is subject to both policies."
          d2="To establish the recommended configuration via GP, set the following UI path to include Guests: Computer Configuration \ Policies \ Windows Settings \ Security Settings \ Local Policies \ User Rights Assignment \ Deny log on locally"

          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=denylog,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=denylog)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=denylog,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=denylog)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
          
     if allowlog == "Failed":
          t1="Local Policies \ User Rights Assignment \ Allow log on through Remote Desktop Services"
          sub="(L1) Ensure 'Allow log on through Remote Desktop Services' is set to 'Administrators, Remote Desktop Users'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="This policy setting determines which users or groups have the right to log on as a Terminal Services client. Remote desktop users require this user right. If your organization uses Remote Assistance as part of its help desk strategy, create a group and assign it this user right through Group Policy."
          d2="To establish the recommended configuration via GP, set the following UI path to Administrators, Remote Desktop Users: Computer Configuration \ Policies \ Windows Settings \ Security Settings \ Local Policies \ User Rights Assignment \ Allow log on through Remote Desktop Services"

          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=allowlog,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=allowlog)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=allowlog,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=allowlog)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if dontdisplay == "Failed":
          t1="Local Policies \ Security Options \ Interactive logon \ Interactive logon: Do not display last user name"
          sub="(L1) Ensure 'Interactive logon: Do not display last user name' is set to 'Enabled'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="This policy setting determines whether the account name of the last user to log on to the client computers in your organization will be displayed in each computer's respective Windows logon screen. Enable this policy setting to prevent intruders from collecting account names visually from the screens of desktop or laptop computers in your organization."
          d2="To establish the recommended configuration via GP, set the following UI path to Enabled: Computer Configuration \ Policies \ Windows Settings \ Security Settings \ Local Policies \ Security Options \ Interactive logon: Do not display last user name"

          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=dontdisplay,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=dontdisplay)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=dontdisplay,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=dontdisplay)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     
     if passexp == "Failed":
          t1="Local Policies \ Security Options \ Interactive logon \ Interactive logon: Prompt user to change password before expiration"
          sub="(L1) Ensure 'Interactive logon: Prompt user to change password before expiration' is set to 'between 5 and 14 days'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="This policy setting determines how far in advance users are warned that their password will expire. It is recommended that you configure this policy setting to at least 5 days but no more than 14 days to sufficiently warn users when their passwords will expire."
          d2="To establish the recommended configuration via GP, set the following UI path to a value between 5 and 14 days: Computer Configuration \ Policies \ Windows Settings \ Security Settings \ Local Policies \ Security Options \ Interactive logon: Prompt user to change password before expiration"

          
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=passexp,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=passexp)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=passexp,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=passexp)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if smartcard == "Failed":
          t1="Local Policies \ Security Options \ Interactive logon \ Interactive logon: Smart card removal behavior"
          sub="(L1) Ensure 'Interactive logon: Smart card removal behavior' is set to 'Lock Workstation' or higher"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="This policy setting determines what happens when the smart card for a logged-on user is removed from the smart card reader."
          d2="To establish the recommended configuration via GP, set the following UI path to Lock Workstation (or, if applicable for your environment, Force Logoff or Disconnect if a Remote Desktop Services session): Computer Configuration \ Policies \ Windows Settings \ Security Settings \ Local Policies \ Security Options \ Interactive logon: Smart card removal behavior"

          
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=smartcard,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=smartcard)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=smartcard,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=smartcard)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if digitaly == "Failed":
          t1="Local Policies \ Security Options \ Microsoft network client \ Microsoft network client: Digitally sign communications (always)"
          sub="(L1) Ensure 'Microsoft network client: Digitally sign communications (always)' is set to 'Enabled'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="This policy setting determines whether packet signing is required by the SMB client component. If you enable this policy setting, the Microsoft network client computer cannot communicate with a Microsoft network server unless that server agrees to sign SMB packets. In mixed environments with legacy client computers, set this option to Disabled because these computers will not be able to authenticate or gain access to domain controllers. "
          d2="To establish the recommended configuration via GP, set the following UI path to Enabled: Computer Configuration \ Policies \ Windows Settings \ Security Settings \ Local Policies \ Security Options \ Microsoft network client: Digitally sign communications (always)"
          
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=digitaly,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=digitaly)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=digitaly,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=digitaly)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if nullsesion == "Failed":
          t1="Local Policies \ Security Options \ Network access \ Network access: Named Pipes that can be accessed anonymously"
          sub="(L1) Ensure 'Network access: Named Pipes that can be accessed anonymously' is set to 'None'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="This policy setting determines which communication sessions, or pipes, will have attributes and permissions that allow anonymous access."
          d2="To establish the recommended configuration via GP, set the following UI path to <blank> (i.e. None): Computer Configuration \ Policies \ Windows Settings \ Security Settings \ Local Policies \ Security Options \ Network access: Named Pipes that can be accessed anonymously"
          
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=nullsesion,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=nullsesion)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=nullsesion,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=nullsesion)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if restricnull == "Failed":
          t1="Local Policies \ Security Options \ Network access \ Network access: Restrict anonymous access to Named Pipes and Shares"
          sub="(L1) Ensure 'Network access: Restrict anonymous access to Named Pipes and Shares' is set to 'Enabled'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="When enabled, this policy setting restricts anonymous access to only those shares and pipes that are named in the Network access: Named pipes that can be accessed anonymously and Network access: Shares that can be accessed anonymously settings. This policy setting controls null session access to shares on your computers by adding RestrictNullSessAccess with the value 1 in the HKEY_LOCAL_MACHINE\System\CurrentControlSet\Services\LanManServer\Parameters registry key. This registry value toggles null session shares on "
          d2="To establish the recommended configuration via GP, set the following UI path to Enabled: Computer Configuration \ Policies \ Windows Settings \ Security Settings \ Local Policies \ Security Options \ Network access: Restrict anonymous access to Named Pipes and Shares"        
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=restricnull,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=restricnull)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=restricnull,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=restricnull)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if scren == "Failed":
          t1="Administrative Templates (User) \ Control Panel \ Personalization \ Enable screen saver"
          sub="(L1) Ensure 'Enable screen saver' is set to 'Enabled'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="This policy setting allows you to manage whether or not screen savers run. If the Screen Saver setting is disabled screen savers do not run and the screen saver section of the Screen Saver tab in Display in Control Panel is disabled. If this setting is enabled a screen saver will run if the following two conditions are met: first, that a valid screen saver is specified on the client via the Screen Saver Executable Name group policy setting or Control Panel on the client. Second, the screensaver timeout is set to a value greater than zero via the Screen Saver Timeout group policy setting or Control Panel on the client."
          d2="To establish the recommended configuration via GP, set the following UI path to Enabled: User Configuration \ Policies \ Administrative Templates \ Control Panel \ Personalization \ Enable screen saver"
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=scren,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=scren)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=scren,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=scren)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if save == "Failed":
          t1="Administrative Templates (User) \ Control Panel \ Personalization \ Force specific screen saver: Screen saver executable name"
          sub="(L1) Ensure 'Force specific screen saver: Screen saver executable name' is set to 'Enabled: scrnsave.scr'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="This policy setting allows you to manage whether or not screen savers run. If the Screen Saver setting is disabled screen savers do not run and the screen saver section of the Screen Saver tab in Display in Control Panel is disabled. If this setting is enabled a screen saver will run if the following two conditions are met: first, that a valid screen saver is specified on the client via the Screen Saver Executable Name group policy setting or Control Panel on the client. Second, the screensaver timeout is set "
          d2="To establish the recommended configuration via GP, set the following UI path to Enabled: scrnsave.scr: User Configuration \ Policies \ Administrative Templates \ Control Panel \ Personalization \ Force specific screen saver"

          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=save,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=save)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=save,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=save)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if screen == "Failed":
          t1="Administrative Templates (User) \ Control Panel \ Personalization \ Password protect the screen saver"
          sub="(L1) Ensure 'Password protect the screen saver' is set to 'Enabled'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="If the Password protect the screen saver setting is enabled, then all screen savers are password protected, if it is disabled then password protection cannot be set on any screen saver."
          d2="To establish the recommended configuration via GP, set the following UI path to Enabled: User Configuration \ Policies \ Administrative Templates \ Control Panel \ Personalization \ Password protect the screen saver"

          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=screen,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=screen)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=screen,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=screen)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if screentime == "Failed":
          t1="Administrative Templates (User) \ Control Panel \ Personalization \ Screen saver timeout"
          sub="(L1) Ensure 'Screen saver timeout' is set to 'Enabled: 900 seconds or fewer, but not 0'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="If the Screen Saver Timeout setting is enabled, then the screen saver will be launched when the specified amount of time has passed since the last user action. Valid values range from 1 to 89,400 seconds (24 hours). The setting has no effect if the wait time is set to zero or no screen saver has been specified."
          d2="To establish the recommended configuration via GP, set the following UI path to Enabled: 900 or fewer, but not 0: User Configuration \ Policies \ Administrative Templates \ Control Panel \ Personalization \ Screen saver timeout"
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=screentime,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=screentime)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=screentime,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=screentime)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if always == "Failed":
          t1="Administrative Templates (User) \ Windows Components \ Windows Installer \ Always install with elevated privileges"
          sub="(L1) Ensure 'Always install with elevated privileges' is set to 'Disabled'"
          nivel="&#9679; Level 1 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 1 + BitLocker"
          d1="Directs Windows Installer to use system permissions when it installs any program on the system. <br /> This setting extends elevated privileges to all programs. These privileges are usually reserved for programs that have been assigned to the user (offered on the desktop), assigned to the computer (installed automatically), or made available in Add or Remove Programs in Control Panel. This setting lets users install programs that require access to directories that the user might not have permission to view or change, including directories on highly restricted computers. "
          d2="To establish the recommended configuration via GP, set the following UI path to Disabled: User Configuration \ Policies \ Administrative Templates \ Windows Components \ Windows Installer \ Always install with elevated privileges"
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=always,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=always)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=always,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=always)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     if prevent == "Failed":
          t1="Administrative Templates (User) \ Windows Components \ Windows Media Player \ Playback \ Prevent Codec Download"
          sub="(L2) Ensure 'Prevent Codec Download' is set to 'Enabled'"
          nivel="&#9679; Level 2 <br /> &nbsp; &nbsp; &nbsp; &nbsp; &#9679; Level 2 + BitLocker"
          d1="This setting controls whether Windows Media Player is allowed to download additional codecs for decoding media files it does not already understand."
          d2="To establish the recommended configuration via GP, set the following UI path to Enabled: User Configuration \ Policies \ Administrative Templates \ Windows Components \ Windows Media Player \ Playback \ Prevent Codec Download"
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=prevent,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=prevent)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     else:
          to=tb_lb_output(titulolb = t1,subtitulolb=sub, descrip1=d1, descrip2=d2)
          tb_lb_output.save(to)
          tp=tb_plugin(severidad=prevent,family='Windows',synopsis=t1,descripcion= d1, solution= d2)
          tb_plugin.save(tp)
          es=tb_estadolb(estadolb=prevent)
          tb_estadolb.save(es)
          his=HistorialEscaneo(cliente_host=nombresistema,dire_ip=IP_host,dire_puerto=' ', vulnerabilidad=sub, fecha_scaneo=fechaactual)
          HistorialEscaneo.save(his)
     
     return redirect('lineabase')
     #return (request)

def cambiarvalor_nuevo_reg1(request):

     key = OpenKey(HKEY_LOCAL_MACHINE, r'Software\Microsoft\Windows\CurrentVersion\Policies\System', 0, KEY_WRITE)
                              #value = (QueryValueEx(key,'DontDisplayLastUserName')[0])  
                              #ya tiene valor 1 se cambia a 0
     #ctypes.windll.shell32.IsUserAnAdmin()
     winreg.SetValueEx(key,'dontdisplaylastusername',0,winreg.REG_DWORD,0)


     return redirect('lineabase')     
  
     CloseKey(key)
 
def cambiartodos(request):
     
          # ejecutar bat
     from subprocess import Popen
     p = Popen("CorregirTodos.bat", cwd=r"C:\Users\Kevin\Documents\Django\ProyectoLineaBase")
     stdout, stderr = p.communicate()  
     print (p)
     
     return redirect('lineabase')

def cambiarvalorMinimunPasswordAge(request):
     from subprocess import Popen
     p = Popen("corregirminumpasswdage.bat", cwd=r"C:\Users\Kevin\Documents\Django\ProyectoLineaBase")
     stdout, stderr = p.communicate()  
     print (p)
     
   
     return redirect('lineabase')

def cambiarvalorminumpasslenght(request):
     from subprocess import Popen
     p = Popen("corregirMinimumPasswordLength.bat", cwd=r"C:\Users\Kevin\Documents\Django\ProyectoLineaBase")
     stdout, stderr = p.communicate()  
     print (p)
        
     return redirect('lineabase')

def cambiarvalorlockooutduration(request):
     from subprocess import Popen
     p = Popen("corregirlockoutduration.bat", cwd=r"C:\Users\Kevin\Documents\Django\ProyectoLineaBase")
     stdout, stderr = p.communicate()  
     print (p)
        
     return redirect('lineabase')
    
def rollbackminimunpassage(request):
     from subprocess import Popen
     p = Popen("rollbackminumpasswdage.bat", cwd=r"C:\Users\Kevin\Documents\Django\ProyectoLineaBase")
     stdout, stderr = p.communicate()  
     print (p)
        
     return redirect('lineabase')
def rollbackminimunpasslenght(request):
     from subprocess import Popen
     p = Popen("rollbackMinimumPasswordLength.bat", cwd=r"C:\Users\Kevin\Documents\Django\ProyectoLineaBase")
     stdout, stderr = p.communicate()  
     print (p)
        
     return redirect('lineabase')
def rollbacklockooutduration(request):
     from subprocess import Popen
     p = Popen("rollbacklockduration.bat", cwd=r"C:\Users\Kevin\Documents\Django\ProyectoLineaBase")
     stdout, stderr = p.communicate()  
     print (p)
     
        
     return redirect('lineabase')




@login_required(login_url='login') 
def vulnerabilidades(request):
     lb= tb_lb_output.objects.all()
     fechaactual = datetime.now() 
     System_Name = platform.uname()[1]
     Os_name=platform.system(),platform.release(),platform.win32_ver()[2] 
     Osversion = platform.version()
     System_Name2 = platform.uname()[2]
     user_name= platform.uname()[1],getpass.getuser()  
    
   
     #buscakb= os.system('cmd /k "systeminfo.exe | findstr KB5014035"')
    # print(buscakb)
          # ejecutar bat
     from subprocess import Popen
     p = Popen("CorregirTodos.bat", cwd=r"C:\Users\Kevin\Documents\Django\ProyectoLineaBase")
     stdout, stderr = p.communicate()  
     print (p)
     
     return redirect('lineabase')
    
    
     
     return render(request,"ProyectoLineaBase\Vulnerabilidades.html",{"historicos": lb,"fecha_actual":fechaactual,"system_name": System_Name,"Os_name":Os_name,"Os_version":Osversion,"system_name2":System_Name2, "usuariologueado":user_name})   

  #consulta
def exportar_a_csv(request):
   #response HttpResponse('Hello')
     response = HttpResponse(content_type ='text/csv')
    # response['Content-Disposition'] ='attachment; filename = pruebacsv.csv'
     
     writer = csv.writer(response)
     
     #encabezzads
     writer.writerow(['severidad','family', 'synopsis','descripcion', 'solution'])
     #personalizada
    
     #recorre lista
     for p in tb_plugin.objects.all()[:10].values_list('severidad','family','synopsis','descripcion','solution'):
          writer.writerow(p)
          
     fecha= datetime.today()
     fechaa = fecha.strftime("%d-%b-%y")
     nombrearchivo= 'LB '+ fechaa +'.xlsx'
     response['Content-Disposition'] ='attachment; filename = "LB ' + fechaa+ '.csv'
     return response 

def exportar_a_excel(request):
     fecha= datetime.today()
     fechaa = fecha.strftime("%d-%b-%y")
     p = HistorialEscaneo.objects.all()
     wb= Workbook()
     ws= wb.active
     ws.title= 'Reporte Linea Base '+ fechaa
     
     ws['B1']= 'Reporte de Linea Base ' + fechaa
     ws['B1'].font= Font(name='Comic Sans MS',size=15,bold=True)
     ws.merge_cells('B1:h1')
     
     ws['b3'].font= Font(name='Comic Sans MS',size=13,bold=True)
     ws['b3']='Cliente Host'
     ws['c3'].font= Font(name='Comic Sans MS',size=13,bold=True)
     ws['c3']='Dire IP'
     ws['d3'].font= Font(name='Comic Sans MS',size=13,bold=True)
     ws['d3']='Vulnerabilidad'
     ws['e3']='Fecha escaneo'
    # ws['f3']='solution'
     
     cont =4
     for pl in p:
          ws.cell(row=cont, column=2).value=pl.cliente_host
          ws.cell(row=cont, column=3).value=pl.dire_ip
          ws.cell(row=cont, column=4).value=pl.vulnerabilidad
          ws.cell(row=cont, column=5).value=str(pl.fecha_scaneo)
          #ws.cell(row=cont, column=6).value=pl.solution
          cont+=1
     #
     
     nombrearchivo= 'LB '+ fechaa +'.xlsx'
     response = HttpResponse(content_type ='application/ms-excel')
     content ='attachment; filename = {0}'.format(nombrearchivo)
     response['Content-Disposition'] =content
     wb.save(response)
     
  
     return response

     
def exportar_a_excel_por_failed(request):
     
     p = tb_plugin.objects.filter(severidad='Failed')
     wb= Workbook()
     ws= wb.active
     ws.title= 'Reporte Linea Base '
     ws['B1']= 'Reporte de Linea Base'
     ws['B1'].font= Font(name='Comic Sans MS',size=13,bold=True)
     ws.merge_cells('B1:h1')
     
     ws['b3']='severidad'
     ws['c3']='family'
     ws['d3']='synopsis'
     ws['e3']='descripcion'
     ws['f3']='solution'
     
     cont =4
     for pl in p:
          ws.cell(row=cont, column=2).value=pl.severidad
          ws.cell(row=cont, column=3).value=pl.family
          ws.cell(row=cont, column=4).value=pl.synopsis
          ws.cell(row=cont, column=5).value=pl.descripcion
          ws.cell(row=cont, column=6).value=pl.solution
          cont+=1
     fecha= datetime.today()
     fechaa = fecha.strftime("%d-%b-%y")
     nombrearchivo= 'LB '+ fechaa +'.xlsx'
     response = HttpResponse(content_type ='application/ms-excel')
     content ='attachment; filename = {0}'.format(nombrearchivo)
     response['Content-Disposition'] =content
     wb.save(response)
     return response

def exportar_a_excel_por_passed(request):
     
     p = tb_plugin.objects.filter(severidad='Passed')
     wb= Workbook()
     ws= wb.active
     ws.title= 'Reporte Linea Base '
     ws['B1']= 'Reporte de Linea Base'
     ws['B1'].font= Font(name='Comic Sans MS',size=13,bold=True)
     ws.merge_cells('B1:h1')
     
     ws['b3']='severidad'
     ws['c3']='family'
     ws['d3']='synopsis'
     ws['e3']='descripcion'
     ws['f3']='solution'
     
     cont =4
     for pl in p:
          ws.cell(row=cont, column=2).value=pl.severidad
          ws.cell(row=cont, column=3).value=pl.family
          ws.cell(row=cont, column=4).value=pl.synopsis
          ws.cell(row=cont, column=5).value=pl.descripcion
          ws.cell(row=cont, column=6).value=pl.solution
          cont+=1
     fecha= datetime.today()
     fechaa = fecha.strftime("%d-%b-%y")
     nombrearchivo= 'LB '+ fechaa +'.xlsx'
     response = HttpResponse(content_type ='application/ms-excel')
     content ='attachment; filename = {0}'.format(nombrearchivo)
     response['Content-Disposition'] =content
     wb.save(response)
     return response

        

@login_required(login_url='login') 
def reporte(request):
     MiniPassAge = MinimumPasswordAge()
     Minpasslength = MinimumPasswordLength()
     MaxPassAge = MaximumPasswordAge()
     PassComplex = PasswordComplexity()
     ClTPass = ClearTextPassword()
     LockDur = LockoutDuration()
       # FIN accountPolicies 
     # Local Policies
     Acccomputer=SeNetworkLogonRight()
     denyaccess= SeDenyNetworkLogonRight()
     denylog = SeDenyInteractiveLogonRight()
     allowlog = SeRemoteInteractiveLogonRight()
   
     #opciones de seguridad
     dontdisplay = DontDisplayLastUserName()
     passexp=PasswordExpiryWARNING()
     smartcard= ScRemoveOption()
     digitaly = RequireSecuritySignature()
     nullsesion=NullSessionPipes()
     restricnull=restrictnullsessaccess()
     lb= tb_lb_output.objects.all()
     
     return render(request,"ProyectoLineaBase\Reporte.html",{"historicos": lb,"minpassle":Minpasslength,"maxpassage":MaxPassAge,"passcomplex":PassComplex,
                                                               "cltpass":ClTPass,"lockdur":LockDur,"acccompu":Acccomputer,"denylog":denylog,
                                                               "denyacce":denyaccess,"allowlo":allowlog,
                                                               "dontdisplays":dontdisplay,"passexperi":passexp,"smartcar":smartcard, 
                                                               "digi":digitaly,"nullse":nullsesion,"restrictnull":restricnull
                                                              })

#login
def registropage(request):
     #validar si ya esta registrado
     if request.user.is_authenticated:
          return redirect('lineabase')
     else:
      
          form = CrearUsuarioForm()   
          
          if request.method =="POST":
               form = CrearUsuarioForm(request.POST)   
               if form.is_valid():
                    form.save()
               user= form.cleaned_data.get('username')
               #mensaje de registro
               messages.success(request,'Cuenta fue creada para ' +user)
               #regresar luego de registrarse redict al login
               return redirect('login')
          
          context = {'form':form}
     return render(request,"ProyectoLineaBase\Registro.html", context)
  
def Loginpage(request):
     if request.user.is_authenticated:
          return redirect('lineabase')
     else:
          if request.method == 'POST':
               username=request.POST.get('username')
               password=request.POST.get('password')
               
               user = authenticate(request,username=username,password=password)
               if user is not None:
                    login(request, user)
                    return redirect('lineabase')          
               else:
                    messages.info(request,'Username or password incorrectos')
               
          context = {}
     
     return render(request,"ProyectoLineaBase\Login.html", context)

def cerrarsesion(request):
     logout(request, )
     return redirect('login')


def Fallidos():
    failed =0
    
    if(SeNetworkLogonRight()=='Failed') ==1 : failed +=1
    if(SeDenyNetworkLogonRight()=='Failed') ==1 : failed+=1
    if(SeDenyInteractiveLogonRight()=='Failed')==1 :failed+=1
    if(SeRemoteInteractiveLogonRight()=='Failed')==1 : failed+=1
    if(DontDisplayLastUserName()=='Failed')==1: failed+=1
    
    if(MinimumPasswordAge()=='Failed')==1: failed+=1
    if(MinimumPasswordLength()=='Failed')==1 : failed+=1
    if(MaximumPasswordAge()=='Failed') ==1 :failed+=1
    if (PasswordComplexity() == 'Failed') ==1 : failed+=1
    if (ClearTextPassword() == 'Failed') ==1 : failed+=1
    if (LockoutDuration() == 'Failed') ==1 : failed+=1
    if (PasswordExpiryWARNING() == 'Failed') ==1 : failed+=1
    if (ScRemoveOption() == 'Failed') ==1 : failed+=1
    if (RequireSecuritySignature() == 'Failed') ==1 :  failed+=1
    if (EnableSecuritySignature() == 'Failed') ==1 : failed+=1
    if (NullSessionPipes() == 'Failed') ==1 : failed+=1
    if (restrictnullsessaccess() == 'Failed') ==1 : failed+=1
    
    if (ScreenSaveActive()=='Failed')==1: failed+=1
    if (SCRNSAVE()=='Failed')==1: failed+=1
    if (ScreenSaverIsSecure()=='Failed')==1: failed+=1
    if (ScreenSaveTimeOut()=='Failed')==1: failed+=1
    if (AlwaysInstallElevated()=='Failed')==1: failed+=1
    if (PreventCodecDownload()=='Failed')==1: failed+=1
   
    return failed
def Passed():
    passed =0 
    if(SeNetworkLogonRight()=='Passed') ==1 : passed +=1
    if(SeDenyNetworkLogonRight()=='Passed') ==1 : passed+=1
    if(SeDenyInteractiveLogonRight()=='Passed')==1 :passed+=1
    if(SeRemoteInteractiveLogonRight()=='Passed')==1 : passed+=1
    if(DontDisplayLastUserName()=='Passed')==1: passed+=1
    
    if(MinimumPasswordAge()=='Passed')==1: passed+=1
    if(MinimumPasswordLength()=='Passed')==1 : passed+=1
    if(MaximumPasswordAge()=='Passed') ==1 :passed+=1
    if (ClearTextPassword() == 'Passed') ==1 : passed+=1
    if (LockoutDuration() == 'Passed') ==1 : passed+=1
    if (PasswordExpiryWARNING() == 'Passed') ==1 : passed+=1
    if (ScRemoveOption() == 'Passed') ==1 : passed+=1
    if (RequireSecuritySignature() == 'Passed') ==1 :  passed+=1
    if (EnableSecuritySignature() == 'Passed') ==1 : passed+=1
    if (NullSessionPipes() == 'Passed') ==1 : passed+=1
    if (restrictnullsessaccess() == 'Passed') ==1 : passed+=1
    
    if (ScreenSaveActive()=='Passed')==1: passed+=1
    if (SCRNSAVE()=='Passed')==1: passed+=1
    if (ScreenSaverIsSecure()=='Passed')==1: passed+=1
    if (ScreenSaveTimeOut()=='Passed')==1: passed+=1
    if (AlwaysInstallElevated()=='Passed')==1: passed+=1
    if (PreventCodecDownload()=='Passed')==1: passed+=1
  
    return passed
#FIn contador
   
# AccountPolicies
op= codecs.open('C:\\Users\\Kevin\\Desktop\\LineaBase WIndows\\Data\\securityoptions.txt','rb','utf-16')
tempacc =op.read()
#1.1.3 Ensure 'Minimum password age' is set to '1 or more day(s)'
def MinimumPasswordAge():
     t = u'MinimunPasswordAge = %s\r\n'
     if t%0 in tempacc:
          return 'Failed'
     for i in range(1,365):
          if t%i in tempacc:
               return 'Passed'

     return 'Failed'    

def MinimumPasswordLength():
	t = u'MinimumPasswordLength = %s\r\n'
	if t%0 in tempacc:
		return 'Failed'
	if t%14 in tempacc:
		return 'Passed'
	return 'Failed'

#1.1.2 Ensure 'Maximum password age' is set to '60 or fewer days, but not 0' 
def MaximumPasswordAge():
	t = u'MaximumPasswordAge = %s\r\n'
	if t%42 in tempacc:
		return 'Failed'
	for i in range(1,91):
		if t%i in tempacc:
			return 'Passed'
	return 'Failed'

#	1.1.5 Ensure 'Password must meet complexity requirements' is set to 'Enabled' 
def PasswordComplexity():
	t = u'PasswordComplexity = %s\r\n'
	if t%0 in tempacc:
		return 'Failed'
	if t%1 in tempacc:
		return 'Passed'
	return 'Failed'

#	1.1.6 Ensure 'Store passwords using reversible encryption' is set to 'Disabled'
def ClearTextPassword():
	t = u'ClearTextPassword = %s\r\n'
	if t%0 in tempacc:
		return 'Passed'
	return 'Failed'
##########	1.2 Account Lockout Policy
#	1.2.1 Ensure 'Account lockout duration' is set to '15 or more minute(s)' -> check
def LockoutDuration():
	t = u'LockoutDuration = %s\r\n'
	for i in range(15,99999):
		if t%i in tempacc:
			return 'Passed'
	if u'LockoutDuration' in tempacc:
		return 'Failed'
	return 'Failed'
# Fin account Policies

#Local Policies
temp=[]
archivo= codecs.open('C:\\Users\\Kevin\\Desktop\\LineaBase WIndows\\Data\\policies.txt','rb','utf-16')
temp =archivo.read()


#	2.2.2 Ensure 'Access this computer from the network' is set to 'Administrators'
def SeNetworkLogonRight():
	if u'SeNetworkLogonRight = *S-1-1-0,*S-1-5-32-544,*S-1-5-32-545,*S-1-5-32-551\r\n' in temp:
		return 'Failed'
	if u'SeNetworkLogonRight = *S-1-5-32-544\r\n' in temp:
		return 'Passed'
	return 'Failed'

#	2.2.16 Ensure 'Deny access to this computer from the network' to include 'Guests, Local account' 
def SeDenyNetworkLogonRight():
	if u'SeDenyNetworkLogonRight = *S-1-5-32-546' in temp:
		return 'Failed'
	if u'SeDenyNetworkLogonRight = %s,*S-1-5-32-546\r\n'%getpass.getuser() in temp:
		return 'Passed'
	return 'Failed'

#	2.2.19 Ensure 'Deny log on locally' to include 'Guests'
def SeDenyInteractiveLogonRight():
	if u'SeDenyInteractiveLogonRight = *S-1-5-32-546\r\n' in temp:
		return 'Passed'
	return 'Failed'
#Allow log on through Remote Desktop Services
def SeRemoteInteractiveLogonRight():
	if u'SeRemoteInteractiveLogonRight = *S-1-5-32-544,*S-1-5-32-555\r\n' in temp:
		return 'Passed'
	return 'Failed'

#Opciones de seguridad regidit

def DontDisplayLastUserName():
	try: 
		key = OpenKey(HKEY_LOCAL_MACHINE, r'Software\Microsoft\Windows\CurrentVersion\Policies\System', 0, KEY_READ)
		value = (QueryValueEx(key,'DontDisplayLastUserName')[0])            
	except:
		return "Failed"
	if value == 0 :
		return "Failed"
	if value == 1 :
		return "Passed"
	return "Failed"
	CloseKey(key)

#		2.3.7.6 Ensure 'Interactive logon: Prompt user to change password before expiration' is set to 'between 5 and 14 days'
def PasswordExpiryWARNING():
     try:
          key= OpenKey(HKEY_LOCAL_MACHINE,r'Software\Microsoft\Windows NT\CurrentVersion\Winlogon',0,KEY_READ)
          value=(QueryValueEx(key,'PasswordExpiryWARNING')[4])
     except:
	#t = u'MACHINE\Software\Microsoft\Windows NT\CurrentVersion\Winlogon\PasswordExpiryWARNING=4,%s\r\n'
          return 'Failed'
     t=5
     if t%'14' in temp:
          return 'Passed'
     if t%'13' in temp or t%'12' in temp or t%'11' in temp or t%'10' in temp or t%'9' in temp or t%'8' in temp or t%'7' in temp or t%'6' in temp or t%'5' in temp:
          return 'Passed'
     return 'Failed'

#		2.3.7.7 Ensure 'Interactive logon: Smart card removal behavior' is set to 'Lock Workstation' or higher
def ScRemoveOption():
	t = u'MACHINE\Software\Microsoft\Windows NT\CurrentVersion\Winlogon\ScRemoveOption=1,"%s"\r\n'
	if t%'0' in temp:
		return 'Failed'
	if t%'1' in temp or t%'2' in temp or t%'3' in temp:
		return 'Passed'
	return 'Failed'

#	2.3.8 Microsoft network client
#		2.3.8.1 Ensure 'Microsoft network client: Digitally sign communications (always)' is set to 'Enabled'
def RequireSecuritySignature():
	try: 
		key = OpenKey(HKEY_LOCAL_MACHINE, r'System\CurrentControlSet\Services\LanmanWorkstation\Parameters', 0, KEY_READ)
		value = (QueryValueEx(key,'RequireSecuritySignature')[0])
	except:
		return 'Failed'
	if value == 0 :
		return 'Failed'
	if value == 1 :
		return 'Passed'
	return 'Failed'
	CloseKey(key)
#		2.3.8.2 Ensure 'Microsoft network client: Digitally sign communications (if server agrees)' is set to 'Enabled'
def EnableSecuritySignature():
	try: 
		key = OpenKey(HKEY_LOCAL_MACHINE, r'System\CurrentControlSet\Services\LanmanWorkstation\Parameters', 0, KEY_READ)
		value = (QueryValueEx(key,'EnableSecuritySignature')[0])
	except:
		return 'Failed'
	if value == 1 :
		return 'Passed'
	return 'Failed'
	CloseKey(key)

#		2.3.10.6 Ensure 'Network access: Named Pipes that can be accessed anonymously' is set to 'None' 
def NullSessionPipes():
	try: 
		key = OpenKey(HKEY_LOCAL_MACHINE, r'SYSTEM\CurrentControlSet\Services\LanManServer\Parameters', 0, KEY_READ)
		value = (QueryValueEx(key,'NullSessionPipes')[0])
	except:
		return 'Failed'
	t = [u'<blank>']
	if (value == t) == 1:
		return 'Passed'
	return 'Failed'
	CloseKey(key)

#		2.3.10.9 Ensure 'Network access: Restrict anonymous access to Named Pipes and Shares' is set to 'Enabled'
def restrictnullsessaccess():
	try: 
		key = OpenKey(HKEY_LOCAL_MACHINE, r'SYSTEM\CurrentControlSet\Services\LanManServer\Parameters', 0, KEY_READ)
		value = (QueryValueEx(key,'restrictnullsessaccess')[0])
	except:
		return 'Failed'
	if value == 1:
		return 'Passed'
	return 'Failed'
	CloseKey(key)
 
 
def get_value_regedit_windows():
    key = OpenKey(HKEY_LOCAL_MACHINE, r'Software\Policies\Microsoft\Windows\Installer')
    for i in range(100):
        try:
             
             n,v,c  = EnumValue(key,i)
             print (i,"-",n,":",v,":") 
            # if i== 15: 
             #print (i,"-",n,":",v,":")
                #return (v)   
               # return (v)
        except:
            break;
    CloseKey(key)


def get_value_regedit_system():
    key = OpenKey(HKEY_LOCAL_MACHINE, r'Software\Microsoft\Windows\CurrentVersion\Policies\System')
    for i in range(100):
        try:
             
             n,v,c  = EnumValue(key,i)
             #print (i,"-",n,":",v,":") 
             if i== 15: 
             #print (i,"-",n,":",v,":")
                #return (v)   
                return (v)
        except:
            break;
    CloseKey(key)
    

##########	19.1 Control Panel
#	19.1.1 Add or Remove Programs
#	19.1.2 Display
#	19.1.3 Personalization
#		19.1.3.1 Ensure 'Enable screen saver' is set to 'Enabled'

def get_user_sid():
	username = "%s"%getpass.getuser()
	b = "wmic useraccount where name='%s' get sid"%username
	sid= os.popen(b).read()
	return sid.split()[1]

def ScreenSaveActive():
	try: 
		key = OpenKey(HKEY_USERS, r'%s\Software\Policies\Microsoft\Windows\Control Panel\Desktop'%get_user_sid(), 0, KEY_ALL_ACCESS)
		value = int(QueryValueEx(key,'ScreenSaveActive')[0])
	except:
		return "Failed"
	if value == 1 :
		return 'Passed'
	return 'Failed';
	CloseKey(key)
	
#		19.1.3.2 Ensure 'Force specific screen saver: Screen saver executable name' is set to 'Enabled: scrnsave.scr'	
def SCRNSAVE():
	try: 
		key = OpenKey(HKEY_USERS, r'%s\Software\Policies\Microsoft\Windows\Control Panel\Desktop'%get_user_sid(), 0, KEY_ALL_ACCESS)
		value = (QueryValueEx(key,'SCRNSAVE.EXE')[0])
	except:
		return "Failed"
	temp = 'scrnsave.scr'
	if (value == temp) == 1 :
		return 'Passed'
	return 'Failed';
	CloseKey(key)
	
#		19.1.3.3 Ensure 'Password protect the screen saver' is set to 'Enabled'
def ScreenSaverIsSecure():
	try: 
		key = OpenKey(HKEY_USERS, r'%s\Software\Policies\Microsoft\Windows\Control Panel\Desktop'%get_user_sid(), 0, KEY_ALL_ACCESS)
		value = int(QueryValueEx(key,'ScreenSaverIsSecure')[0])
	except:
		return "Failed"
	if value == 1 :
		return 'Passed'
	return 'Failed';
	CloseKey(key)
	
#		19.1.3.4 Ensure 'Screen saver timeout' is set to 'Enabled: 900 seconds or fewer, but not 0'
def ScreenSaveTimeOut():
	try: 
		key = OpenKey(HKEY_USERS, r'%s\Software\Policies\Microsoft\Windows\Control Panel\Desktop'%get_user_sid(), 0, KEY_ALL_ACCESS)
		value = int(QueryValueEx(key,'ScreenSaveTimeOut')[0])
	except:
		return "Failed"
	if value <= 900 and value > 0 :
		return 'Passed'
	elif value == 0:
		return "Failed"
	return 'Failed';
	CloseKey(key)
 
def AlwaysInstallElevated():
	try: 
		key = OpenKey(HKEY_USERS, r'%s\Software\Policies\Microsoft\Windows\Installer'%get_user_sid(), 0, KEY_ALL_ACCESS)
		value = (QueryValueEx(key,'AlwaysInstallElevated')[0])
	except:
		return "Failed"
	if value == 0:
		return 'Passed'
	return 'Failed';
	CloseKey(key)

#		19.7.41.2.1 Ensure 'Prevent Codec Download' is set to 'Enabled'
def PreventCodecDownload():
	try: 
		key = OpenKey(HKEY_USERS, r'%s\Software\Policies\Microsoft\WindowsMediaPlayer'%get_user_sid(), 0, KEY_ALL_ACCESS)
		value = (QueryValueEx(key,'PreventCodecDownload')[0])
	except:
		return "Failed"
	if value == 1:
		return 'Passed'
	return 'Failed';
	CloseKey(key)
	